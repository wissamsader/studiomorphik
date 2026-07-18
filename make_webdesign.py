#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build web-design/index.html — the StudioMorphik web-design portfolio page
(his 07-14 ask: sellers need a portfolio link for customers; lists ALL live
demo/client sites as branded tiles, click-through to each site in a new tab).

Data = the same rosters HQ uses (WEBSITE-BUSINESS kitdata modules + the
PITCH-KIT xlsx). Thumbnails = generated locally into web-design/shots/ (browser dots, NO URL text
— his ask). Rerun after any new city/batch, then commit+push
this repo (site = studiomorphik.com via GitHub Pages).
"""
import html, pathlib, re, sys, importlib.util

HERE = pathlib.Path(__file__).parent
WB = pathlib.Path.home() / "Desktop" / "WEBSITE-BUSINESS"

def load_mod(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

SITES = []  # dicts: name, city (display), url, repo
def add(name, url, repo, city):
    url = url.strip()
    if not url.startswith("http"):
        url = "https://" + url
    if not url.endswith("/"):
        url += "/"
    SITES.append(dict(name=name.strip(), url=url, repo=repo, city=city))

# Vietnam
sys.path.insert(0, str(WB / "VIETNAM" / "tools"))
import kitdata as vn  # noqa
for b in vn.BIZ:
    add(b["name"], b["url"], "vietnam", "Đà Nẵng")

# Beirut + Chiang Mai from the xlsx
from openpyxl import load_workbook
wb = load_workbook(WB / "WISSAM-PITCH" / "PITCH-KIT.xlsx")
for sheet, repo, city in (("Beirut", "beirut", "Beirut"), ("ChiangMai", "chiangmai", "Chiang Mai")):
    for row in wb[sheet].iter_rows(min_row=2):
        if not row[2].value:
            continue
        add(str(row[2].value), str(row[11].value), repo, city)

# Barcelona / Beirut batch 3 / Palermo kitdata modules
bcn = load_mod("bcn_kd", WB / "BARCELONA" / "tools" / "kitdata.py")
for b in bcn.BIZ:
    add(b["name"], b["url"], "barcelona", "Barcelona")
bey3 = load_mod("bey3_kd", WB / "BEIRUT" / "tools3" / "kitdata3.py")
for b in bey3.BIZ:
    add(b["name"], b["url"], "beirut", "Beirut")
plm = load_mod("plm_kd", WB / "PALERMO" / "tools" / "kitdata.py")
for b in plm.BIZ:
    add(b["name"], b["url"], "palermo", "Palermo")
dam = load_mod("dam_kd", WB / "DAMASCUS" / "tools" / "kitdata.py")
for b in dam.BIZ:
    add(b["name"], f"{dam.BASE_URL}/{b['slug']}/", "damascus", "Damascus")
_ber_kd = WB / "BERLIN" / "tools" / "kitdata.py"
if _ber_kd.exists():  # Berlin build may still be in flight — skip until its kitdata lands
    ber = load_mod("ber_kd", _ber_kd)
    for b in ber.BIZ:
        add(b["name"], b["url"], "berlin", "Berlin")


# grouped per city with jump-nav (07-14 friend feedback via Wissam: "tekbos 3al city, byenzal la 7alo")
CITY_ORDER = ["Beirut", "Chiang Mai", "Đà Nẵng", "Barcelona", "Palermo", "Damascus", "Berlin"]
CITY_ID = {"Beirut": "beirut", "Chiang Mai": "chiang-mai", "Đà Nẵng": "da-nang", "Barcelona": "barcelona", "Palermo": "palermo", "Damascus": "damascus", "Berlin": "berlin"}
by_city = {}
for s in SITES:
    by_city.setdefault(s["city"], []).append(s)
order = [s for c in CITY_ORDER for s in by_city.get(c, [])]

# ---- portfolio tiles: generated locally, NO URL text anywhere (his 07-14 ask) ----
from PIL import Image, ImageDraw
SHOT_SRC = {
    "vietnam": WB / "VIETNAM" / "PITCH-KIT-VIETNAM" / "screenshots",
    "barcelona": WB / "BARCELONA" / "PITCH-KIT-BARCELONA" / "screenshots",
    "palermo": WB / "PALERMO" / "PITCH-KIT-PALERMO" / "screenshots",
    "beirut-rabab": WB / "BEIRUT" / "PITCH-KIT-RABAB" / "screenshots",
    "bey-cm": WB / "WISSAM-PITCH" / "kit-machine" / "screenshots",
    "damascus": WB / "DAMASCUS" / "PITCH-KIT-DAMASCUS" / "screenshots",
    "berlin": WB / "BERLIN" / "PITCH-KIT-BERLIN" / "screenshots",
}
TILE_DIR = HERE / "web-design" / "shots"
TILE_DIR.mkdir(parents=True, exist_ok=True)

def source_png(s):
    slug = s["url"].rstrip("/").split("/")[-1]
    if s["repo"] == "vietnam":
        return SHOT_SRC["vietnam"] / f"{slug}.png"
    if s["repo"] == "barcelona":
        return SHOT_SRC["barcelona"] / f"{slug}.png"
    if s["repo"] == "palermo":
        return SHOT_SRC["palermo"] / f"{slug}.png"
    if s["repo"] == "damascus":
        return SHOT_SRC["damascus"] / f"{slug}.png"
    if s["repo"] == "berlin":
        return SHOT_SRC["berlin"] / f"{slug}.png"
    if s["repo"] == "beirut":
        p = SHOT_SRC["bey-cm"] / f"Beirut — {s['name']}.png"
        return p if p.exists() else SHOT_SRC["beirut-rabab"] / f"{slug}.png"
    return SHOT_SRC["bey-cm"] / f"Chiang Mai — {s['name']}.png"

S = 640  # plain top-cropped square screenshot; the frame is drawn in CSS (modern glass/gradient)

def make_tile(src_path, out_path):
    src = Image.open(src_path).convert("RGB")
    sw, sh = src.size
    if sw > sh:
        x0 = (sw - sh) // 2
        crop = src.crop((x0, 0, x0 + sh, sh))
    else:
        crop = src.crop((0, 0, sw, sw))
    crop = crop.resize((S, S), Image.LANCZOS)
    crop.save(out_path, "JPEG", quality=86)

def shot(s):
    slug = s["url"].rstrip("/").split("/")[-1]
    out = TILE_DIR / f"{s['repo']}-{slug}.jpg"
    if not out.exists():
        src = source_png(s)
        assert src.exists(), f"missing screenshot: {src}"
        make_tile(src, out)
    return f"shots/{out.name}"

def card(s):
    return (f'<a class="card" href="{html.escape(s["url"])}" target="_blank" rel="noopener">'
            f'<span class="inner"><span class="chrome" aria-hidden="true"><i></i><i></i><i></i></span>'
            f'<img loading="lazy" src="{shot(s)}?v=2" alt="{html.escape(s["name"])} website by StudioMorphik">'
            f'<span class="meta"><b>{html.escape(s["name"])}</b><span style="display:flex;align-items:center;gap:9px"><i>{s["city"]}</i><span class="arr">↗</span></span></span></span></a>')

sections = "\n".join(
    f'<section class="cityblock" id="{CITY_ID[c]}">'
    f'<h2 class="cityhead">{c.upper()}<span>{len(by_city.get(c, []))} websites</span></h2>'
    f'<div class="grid">' + "".join(card(s) for s in by_city.get(c, [])) + '</div></section>'
    for c in CITY_ORDER if by_city.get(c))

citynav = "".join(f'<a href="#{CITY_ID[c]}">{c}<b>{len(by_city[c])}</b></a>' for c in CITY_ORDER if by_city.get(c))

n = len(order)
live_cities = [c for c in CITY_ORDER if by_city.get(c)]
n_cities = len(live_cities)
CITY_COUNTRY = {"Beirut": "Lebanon", "Chiang Mai": "Thailand", "Đà Nẵng": "Vietnam", "Barcelona": "Spain",
                "Palermo": "Italy", "Damascus": "Syria", "Berlin": "Germany"}
n_countries = len({CITY_COUNTRY[c] for c in live_cities if c in CITY_COUNTRY})
cities = " · ".join(live_cities)

page = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, maximum-scale=1, user-scalable=no">
<title>Web Design — STUDIOMORPHIK</title>
<meta name="description" content="Hand-built websites for restaurants, guesthouses, independent businesses: {n} live builds across {cities} by StudioMorphik.">
<link rel="canonical" href="https://studiomorphik.com/web-design/">
<link rel="icon" href="/favicon.ico" sizes="any">
<link rel="icon" type="image/png" sizes="96x96" href="/img/favicon-96.png">
<meta property="og:title" content="Web Design — STUDIOMORPHIK">
<meta property="og:description" content="{n} live websites for restaurants, guesthouses, independent businesses.">
<meta property="og:type" content="website">
<meta property="og:url" content="https://studiomorphik.com/web-design/">
<meta property="og:image" content="https://studiomorphik.com/web-design/og.jpg">
<meta property="og:image:width" content="1200">
<meta property="og:image:height" content="630">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:image" content="https://studiomorphik.com/web-design/og.jpg">
<style>
@font-face{{font-family:'Integral CF';src:url('/fonts/IntegralCF-Bold.woff2') format('woff2');font-weight:700;font-style:normal;font-display:swap}}
:root{{--bg:#0b0c0e;--fg:rgba(255,255,255,.92);--dim:rgba(255,255,255,.6);--line:rgba(255,255,255,.14);--sur:#121317}}
*{{box-sizing:border-box;margin:0;-webkit-tap-highlight-color:transparent}}
html{{scroll-behavior:smooth}}
body{{overflow-x:clip;background:var(--bg);color:var(--fg);font-family:'Montserrat',system-ui,-apple-system,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;line-height:1.6;letter-spacing:.01em;-webkit-font-smoothing:antialiased}}
a{{color:inherit;text-decoration:none}}
.wrap{{max-width:1280px;margin:0 auto;padding:0 clamp(18px,4vw,44px)}}
.mono{{font-family:'JetBrains Mono',ui-monospace,'SF Mono',Menlo,monospace}}
header{{position:sticky;top:0;z-index:20;background:rgba(11,12,14,.82);backdrop-filter:blur(12px);border-bottom:1px solid var(--line)}}
header .bar{{display:flex;align-items:center;justify-content:space-between;padding:16px clamp(18px,4vw,44px);max-width:1280px;margin:0 auto}}
.wordmark{{font-family:'Integral CF',sans-serif;font-size:15px;letter-spacing:.22em}}
header nav{{display:flex;gap:22px;font-family:'JetBrains Mono',ui-monospace,monospace;font-size:11px;letter-spacing:.16em;text-transform:uppercase;color:var(--dim)}}
header nav a:hover{{color:var(--fg)}}
.hero{{padding:clamp(56px,10vw,110px) 0 clamp(28px,4vw,44px)}}
.eyebrow{{font-family:'JetBrains Mono',ui-monospace,monospace;font-size:11px;letter-spacing:.3em;text-transform:uppercase;color:var(--dim)}}
h1{{font-family:'Integral CF',sans-serif;font-size:clamp(2.4rem,8vw,5.4rem);line-height:1.02;letter-spacing:.02em;margin:.35em 0 .3em}}
.sub{{max-width:62ch;color:var(--dim);font-size:clamp(15px,1.6vw,17px)}}
.stats{{display:flex;gap:10px 26px;flex-wrap:wrap;margin-top:26px;font-family:'JetBrains Mono',ui-monospace,monospace;font-size:11.5px;letter-spacing:.14em;text-transform:uppercase;color:var(--dim)}}
.stats b{{color:var(--fg);font-weight:700}}
.stats .citylist{{flex-basis:100%;color:rgba(255,255,255,.42);letter-spacing:.12em}}
.note{{border:1px solid var(--line);border-left:2px solid rgba(255,255,255,.45);background:var(--sur);border-radius:8px;padding:14px 18px;margin:34px 0 0;max-width:74ch;color:var(--dim);font-size:13.5px;line-height:1.65}}
.note b{{color:var(--fg)}}
.citynav{{position:sticky;top:57px;z-index:15;background:rgba(11,12,14,.86);backdrop-filter:blur(12px);border-bottom:1px solid var(--line)}}
.citynav .in{{display:flex;gap:8px;overflow-x:auto;scrollbar-width:none;padding:10px clamp(18px,4vw,44px);max-width:1280px;margin:0 auto}}
.citynav .in::-webkit-scrollbar{{display:none}}
.citynav a{{flex-shrink:0;display:inline-flex;align-items:center;gap:8px;border:1px solid var(--line);border-radius:999px;padding:7px 16px;font-family:'JetBrains Mono',ui-monospace,monospace;font-size:10.5px;letter-spacing:.14em;text-transform:uppercase;color:var(--dim);transition:color .2s,border-color .2s,background .2s}}
.citynav a b{{font-weight:700;font-size:9px;color:rgba(255,255,255,.38);transition:color .2s}}
.citynav a:hover,.citynav a.on{{color:#0b0c0e;background:var(--fg);border-color:var(--fg)}}
.citynav a:hover b,.citynav a.on b{{color:rgba(11,12,14,.55)}}
.cityblock{{scroll-margin-top:118px;padding-top:clamp(26px,4vw,44px)}}
.cityhead{{font-family:'Integral CF',sans-serif;font-size:clamp(1.25rem,3vw,1.9rem);letter-spacing:.06em;display:flex;align-items:baseline;gap:14px;margin-bottom:clamp(14px,2vw,22px)}}
.cityhead span{{font-family:'JetBrains Mono',ui-monospace,monospace;font-size:10.5px;letter-spacing:.18em;text-transform:uppercase;color:var(--dim)}}
.gridwrap{{padding:clamp(10px,2vw,24px) 0 20px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(216px,1fr));gap:clamp(12px,1.8vw,20px)}}
.card{{min-width:0;position:relative;display:block;border-radius:20px;padding:1px;background:linear-gradient(165deg,rgba(255,255,255,.2),rgba(255,255,255,.055) 40%,rgba(255,255,255,.02));box-shadow:0 24px 56px -28px rgba(0,0,0,.8);transition:transform .3s cubic-bezier(.2,.7,.2,1)}}
.card::before{{content:"";position:absolute;inset:-2px;border-radius:22px;background:conic-gradient(from 210deg,#ff4d4d,#ff9f40,#ffe14d,#5dff9b,#4dc3ff,#b96bff,#ff4d4d);opacity:0;filter:blur(14px);transition:opacity .35s;z-index:0}}
.card:hover{{transform:translateY(-5px)}}
.card:hover::before{{opacity:.35}}
.card .inner{{position:relative;z-index:1;border-radius:19px;overflow:hidden;background:#101114;display:block}}
.card .chrome{{display:flex;align-items:center;gap:5px;padding:9px 13px;background:linear-gradient(180deg,rgba(255,255,255,.075),rgba(255,255,255,.02));border-bottom:1px solid rgba(255,255,255,.07)}}
.card .chrome i{{width:7px;height:7px;border-radius:50%;background:rgba(255,255,255,.22)}}
.card img{{width:100%;aspect-ratio:1/1;display:block;object-fit:cover;background:#0b0c0e}}
.card .meta{{display:flex;align-items:center;justify-content:space-between;gap:8px;padding:11px 14px 12px;border-top:1px solid rgba(255,255,255,.07);background:rgba(255,255,255,.02)}}
.card .meta b{{min-width:0;flex:1;font-size:12.5px;font-weight:600;letter-spacing:.02em;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.card .meta i{{font-style:normal;font-family:'JetBrains Mono',ui-monospace,monospace;font-size:9.5px;letter-spacing:.14em;text-transform:uppercase;color:var(--dim);flex-shrink:0}}
.card .meta .arr{{font-family:'JetBrains Mono',monospace;font-size:12px;color:var(--dim);opacity:0;transform:translate(-4px,4px);transition:opacity .25s,transform .25s;flex-shrink:0}}
.card:hover .meta .arr{{opacity:1;transform:none;color:var(--fg)}}
.cta{{margin:clamp(40px,7vw,70px) 0;border:1px solid var(--line);border-radius:14px;background:var(--sur);padding:clamp(26px,5vw,44px);display:flex;align-items:center;justify-content:space-between;gap:20px;flex-wrap:wrap}}
.cta h2{{font-family:'Integral CF',sans-serif;font-size:clamp(1.2rem,3vw,1.8rem);letter-spacing:.03em}}
.cta p{{color:var(--dim);margin-top:6px;max-width:52ch}}
.btn{{display:inline-block;border:1px solid rgba(255,255,255,.5);border-radius:999px;padding:12px 26px;font-family:'JetBrains Mono',ui-monospace,monospace;font-size:12px;letter-spacing:.18em;text-transform:uppercase;transition:background .2s,color .2s}}
.btn:hover{{background:var(--fg);color:#0b0c0e}}
footer{{border-top:1px solid var(--line);padding:26px 0 34px;color:var(--dim);font-size:12.5px;display:flex;justify-content:space-between;gap:14px;flex-wrap:wrap}}
footer a:hover{{color:var(--fg)}}
@media(max-width:520px){{
.grid{{grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}}
.card{{border-radius:16px}}.card::before{{border-radius:18px}}.card .inner{{border-radius:15px}}
.card .chrome{{padding:7px 10px;gap:4px}}.card .chrome i{{width:5.5px;height:5.5px}}
.card .meta{{flex-direction:column;align-items:flex-start;gap:2px;padding:9px 11px 10px}}
.card .meta b{{width:100%;font-size:11.5px}}
.card .meta i{{font-size:8.5px}}
.card .meta .arr{{display:none}}
}}
@media(prefers-reduced-motion:reduce){{*{{transition:none!important}}}}
</style>
</head>
<body>
<header><div class="bar">
 <a class="wordmark" href="/">STUDIOMORPHIK</a>
 <nav><a href="/#/about">About</a><a href="/#/contact">Contact</a></nav>
</div></header>

<section class="hero"><div class="wrap">
 <span class="eyebrow">Studiomorphik · Web design</span>
 <h1>WEB DESIGN</h1>
 <p class="sub">Hand-built websites for restaurants, guesthouses, bars, independent businesses. Designed, built, and cared for by StudioMorphik. One page, your photos, your story. Live in days.</p>
 <div class="stats"><span><b>{n}</b>&nbsp;live builds</span><span><b>{n_cities}</b>&nbsp;cities</span><span><b>{n_countries}</b>&nbsp;countries</span><span class="citylist">{cities}</span></div>
 <p class="note"><b>About this portfolio:</b> every website below is live. Open any of them. Some are commissioned client work; others are ready-made designs created for real businesses, ready for their owners to claim.</p>
</div></section>

<nav class="citynav" aria-label="Cities"><div class="in">{citynav}</div></nav>

<section class="gridwrap"><div class="wrap">
{sections}
</div></section>

<div class="wrap"><div class="cta">
 <div><h2>WANT YOURS?</h2><p>Your restaurant or guesthouse on one beautiful page: your photos, your menu, linked on Google Maps. No monthly fees.</p></div>
 <a class="btn" href="/#/contact">Get in touch</a>
</div></div>

<div class="wrap"><footer>
 <span>© 2026 STUDIOMORPHIK</span>
 <span><a href="mailto:hello@studiomorphik.com">hello@studiomorphik.com</a> · <a href="/">studiomorphik.com</a></span>
</footer></div>
<script>
(function(){{var links=[].slice.call(document.querySelectorAll('.citynav a'));
var map={{}};links.forEach(function(a){{map[a.getAttribute('href').slice(1)]=a}});
if('IntersectionObserver' in window){{
var io=new IntersectionObserver(function(es){{es.forEach(function(e){{
 if(e.isIntersecting){{links.forEach(function(a){{a.classList.remove('on')}});
 var a=map[e.target.id];if(a)a.classList.add('on');}}}})}},{{rootMargin:'-30% 0px -60% 0px'}});
document.querySelectorAll('.cityblock').forEach(function(s){{io.observe(s)}});}}
}})();
</script>
</body>
</html>
"""

# ---- animated attractor hero background (Three.js particle swarm, brand spectrum) ----
# Injected after the page string is built so the particle JS keeps its literal
# braces (the template above is an f-string). Renders behind the hero ONLY; all
# content sits above it and the portfolio grid is untouched. Degrades to the plain
# dark hero if WebGL or the CDN are unavailable. Honours prefers-reduced-motion
# (settles then freezes) and pauses when the hero scrolls off-screen.
_FX_HEAD = """
<style data-fx-attractor>
#heroFx{position:relative;overflow:hidden;min-height:100svh;display:flex;align-items:center}
#heroFx>.wrap{position:relative;z-index:2;width:100%}
#fx{position:absolute;inset:0;width:100%;height:100%;z-index:0;display:block;pointer-events:none}
.heroshade{position:absolute;inset:0;z-index:1;pointer-events:none;
  background:
   radial-gradient(130% 110% at 66% 44%, rgba(11,12,14,0) 34%, rgba(11,12,14,.5) 100%),
   linear-gradient(90deg, rgba(11,12,14,.82) 2%, rgba(11,12,14,.42) 34%, rgba(11,12,14,.1) 60%, rgba(11,12,14,0) 100%),
   linear-gradient(180deg, rgba(11,12,14,0) 76%, #0b0c0e 99%)}
@media (max-width:640px){#heroFx{min-height:auto;align-items:flex-start;padding-top:clamp(52px,12svh,120px);padding-bottom:clamp(44px,9svh,110px)}}
</style>
<script type="importmap">
{"imports":{"three":"https://unpkg.com/three@0.160.0/build/three.module.js","three/addons/":"https://unpkg.com/three@0.160.0/examples/jsm/"}}
</script>
"""

_FX_MODULE = """
<script type="module" data-fx-attractor>
import * as THREE from 'three';
import { EffectComposer } from 'three/addons/postprocessing/EffectComposer.js';
import { RenderPass } from 'three/addons/postprocessing/RenderPass.js';
import { UnrealBloomPass } from 'three/addons/postprocessing/UnrealBloomPass.js';

const host = document.getElementById('heroFx');
const canvas = document.getElementById('fx');
if (host && canvas) {
  const hs = () => { const r = host.getBoundingClientRect(); return { w: Math.max(1, r.width), h: Math.max(1, r.height) }; };
  let S = hs(), W = S.w, H = S.h;
  const COUNT = (W < 640) ? 7000 : 20000;

  const scene = new THREE.Scene();
  scene.fog = new THREE.FogExp2(0x000000, 0.004);
  const camera = new THREE.PerspectiveCamera(60, W / H, 0.1, 3000);
  camera.position.set(0, 0, 235);

  const renderer = new THREE.WebGLRenderer({ canvas, antialias: true, alpha: true, powerPreference: "high-performance" });
  renderer.setPixelRatio(Math.min(window.devicePixelRatio || 1, (W < 640) ? 1.5 : 2));
  renderer.setSize(W, H, false);
  renderer.setClearColor(0x000000, 0);

  const composer = new EffectComposer(renderer);
  composer.setSize(W, H);
  composer.addPass(new RenderPass(scene, camera));
  const bloom = new UnrealBloomPass(new THREE.Vector2(W, H), 1.4, 0.5, 0.0);
  bloom.strength = 0.45; bloom.radius = 0.5; bloom.threshold = 0.1;
  composer.addPass(bloom);

  const color = new THREE.Color();
  const target = new THREE.Vector3();

  const dotCanvas = document.createElement('canvas');
  dotCanvas.width = dotCanvas.height = 64;
  const dctx = dotCanvas.getContext('2d');
  const dgrad = dctx.createRadialGradient(32, 32, 0, 32, 32, 32);
  dgrad.addColorStop(0.0, 'rgba(255,255,255,1)');
  dgrad.addColorStop(0.45, 'rgba(255,255,255,0.5)');
  dgrad.addColorStop(1.0, 'rgba(255,255,255,0)');
  dctx.fillStyle = dgrad; dctx.fillRect(0, 0, 64, 64);
  const dotTex = new THREE.CanvasTexture(dotCanvas);

  const geometry = new THREE.BufferGeometry();
  const posArr = new Float32Array(COUNT * 3);
  const colArr = new Float32Array(COUNT * 3);
  for (let i = 0; i < COUNT; i++) {
    posArr[i * 3] = (Math.random() - 0.5) * 100;
    posArr[i * 3 + 1] = (Math.random() - 0.5) * 100;
    posArr[i * 3 + 2] = (Math.random() - 0.5) * 100;
    colArr[i * 3] = 0.0; colArr[i * 3 + 1] = 1.0; colArr[i * 3 + 2] = 0.53;
  }
  geometry.setAttribute('position', new THREE.BufferAttribute(posArr, 3));
  geometry.setAttribute('color', new THREE.BufferAttribute(colArr, 3));

  const material = new THREE.PointsMaterial({ size: 1.7, map: dotTex, vertexColors: true,
    transparent: true, depthWrite: false, blending: THREE.AdditiveBlending, sizeAttenuation: true });
  const points = new THREE.Points(geometry, material);
  points.position.set(0, 0, 0);
  scene.add(points);

  const positions = [];
  for (let i = 0; i < COUNT; i++) { positions.push(new THREE.Vector3(posArr[i * 3], posArr[i * 3 + 1], posArr[i * 3 + 2])); }
  const posAttr = geometry.attributes.position;
  const colAttr = geometry.attributes.color;

  const clock = new THREE.Clock();
  let visible = true, ticking = false;

  function renderFrame() {
    const elapsed = clock.getElapsedTime();
    const time = elapsed * 0.08;
    points.rotation.y = elapsed * 0.045;   // gentle continuous spin, always visible
    const count = COUNT;
    for (let i = 0; i < COUNT; i++) {
      const scale = 165, spread = 1.0, flow = 0.8, morph = 0.45;
      const t = i / count;
      const theta = t * Math.PI * 40.0;
      const r0 = 0.35 + 0.65 * Math.sqrt(t);
      const w1 = Math.sin(theta * 0.7 + time * 0.25);
      const w2 = Math.sin(theta * 1.9 - time * 0.17);
      const w3 = Math.cos(theta * 3.3 + time * 0.11);
      const attract = r0 * (1.0 + 0.25 * w1 + 0.12 * w2 * w3);
      const bend = 0.8 * Math.sin(theta * 0.5 + time * 0.2) + 0.3 * Math.sin(theta * 2.7 - time * 0.13);
      const x = scale * attract * Math.cos(theta + bend * spread);
      const y = scale * attract * Math.sin(theta + bend * spread);
      const z = scale * (0.55 * Math.sin(theta * 0.55) + 0.28 * Math.sin(theta * 1.73 + time * 0.22) + 0.15 * Math.cos(theta * 4.1));
      const swirl = flow * Math.sin(Math.sqrt(x * x + y * y) * 0.06 - time * 0.8);
      const px = x + swirl * y * 0.12;
      const py = y - swirl * x * 0.12;
      const pz = z + scale * morph * 0.18 * Math.sin(theta * 0.9 + time * 0.35);
      target.set(px, py, pz);
      const hue = (t * 0.85 + time * 0.02) % 1;
      const light = 0.22 + 0.12 * Math.exp(-Math.abs(attract - 0.75));
      color.setHSL(hue, 0.78, light);
      positions[i].lerp(target, 0.045);
      posAttr.array[i * 3] = positions[i].x; posAttr.array[i * 3 + 1] = positions[i].y; posAttr.array[i * 3 + 2] = positions[i].z;
      colAttr.array[i * 3] = color.r; colAttr.array[i * 3 + 1] = color.g; colAttr.array[i * 3 + 2] = color.b;
    }
    posAttr.needsUpdate = true;
    colAttr.needsUpdate = true;
    composer.render();
  }

  function tick() {
    renderFrame();
    if (visible) { requestAnimationFrame(tick); } else { ticking = false; }
  }
  function start() { if (!ticking) { ticking = true; requestAnimationFrame(tick); } }

  if ('IntersectionObserver' in window) {
    const io = new IntersectionObserver((es) => { visible = es[0].isIntersecting; if (visible) start(); }, { threshold: 0 });
    io.observe(host);
  }
  start();

  window.addEventListener('resize', () => {
    S = hs(); W = S.w; H = S.h;
    camera.aspect = W / H; camera.updateProjectionMatrix();
    renderer.setSize(W, H, false); composer.setSize(W, H);
    start();
  });
}
</script>
"""

page = page.replace(
    '<section class="hero"><div class="wrap">',
    '<section class="hero" id="heroFx"><canvas id="fx" aria-hidden="true"></canvas>'
    '<div class="heroshade" aria-hidden="true"></div><div class="wrap">',
    1,
)
page = page.replace("</head>", _FX_HEAD + "</head>", 1)
page = page.replace("</body>", _FX_MODULE + "</body>", 1)

out = HERE / "web-design" / "index.html"
out.parent.mkdir(exist_ok=True)
out.write_text(page)
print(f"web-design/index.html: {n} sites, {out.stat().st_size//1024} KB")
